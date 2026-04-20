"""
Automated paper selection for systematic literature review.
Focus: Decision Support in High-Mix Low-Volume (HMLV) Manufacturing
Emphasis: Industry 4.0 technologies such as AI, Digital Twin, IoT, Optimization, Simulation

Output: Excel sheet with key_id and Relevance Score columns (++, +, *, -)
"""

import pandas as pd
import json
import re
from pathlib import Path
from typing import Tuple


def load_paths_config():
    for parent in Path(__file__).resolve().parents:
        config_path = parent / "paths.json"
        if config_path.exists():
            with open(config_path, "r", encoding="utf-8") as f:
                return json.load(f), config_path.parent
    raise FileNotFoundError("Could not find paths.json from script location")


def resolve_from_root(root_dir: Path, relative_path: str) -> str:
    return str((root_dir / relative_path.lstrip("./")).resolve())


class PaperSelector:
    """
    Intelligent paper selection based on relevance to Decision Support in HMLV Manufacturing.

    Scoring criteria:
    - ++ (Most Relevant): Strong technical focus on decision support using Industry 4.0 in manufacturing
    - + (Relevant): Optimization/simulation/scheduling in production context with some digital focus
    - * (Maybe): General digitalization or technology applied in manufacturing, no decision support
    - - (Exclude): Non-manufacturing or purely business/managerial/policy focus
    """

    def __init__(self):
        # --- Core Industry 4.0 and AI-related terms ---
        self.tech_core = re.compile(
            r'\b('
            r'machine\s*learning|deep\s*learning|artificial\s*intelligence|neural\s*network|reinforcement\s*learning|'
            r'digital\s*twin|cyber[-\s]*physical[-\s]*system|cps|cyber[-\s]*physical[-\s]*production[-\s]*systems?|cpps|'
            r'industrial\s*iot|iiot|internet\s+of\s+things|iot|'
            r'edge\s*computing|embedded\s*ai|analytics|predictive\s*maintenance|'
            r'fault\s*detection|data[-\s]*driven|'
            r'robot|robotics|robotic\s*system|'
            r'human[-\s]*robot\s*collaboration|hrc|'
            r'cobot|cobots|'
            r'autonomous\s*vehicle|agv|agvs|'
            r'autonomous\s*mobile\s*robot|amr'
            #r'cloud\s*computing|big\s*data|'
            r')\b',
            re.IGNORECASE
        )


        # --- Methods used in AI/ML/Digital Twin/Optimization ---
        self.methods_core = re.compile(
            r'\b('
            r'cnn|convolutional\s*neural\s*network|rnn|recurrent\s*neural\s*network|lstm|gru|'
            r'transformer|gpt|bert|autoencoder|variational\s*autoencoder|vae|gan|'
            r'generative\s*adversarial\s*network|diffusion\s*model|graph\s*neural\s*network|gnn|'
            r'siamese\s*network|attention\s*mechanism|vision\s*transformer|vit|mlp|'
            r'multi[-\s]*layer\s*perceptron|'
            r'supervised\s*learning|unsupervised\s*learning|semi[-\s]*supervised\s*learning|'
            r'self[-\s]*supervised\s*learning|transfer\s*learning|federated\s*learning|'
            r'representation\s*learning|metric\s*learning|ensemble\s*learning|'
            r'multi[-\s]*task\s*learning|few[-\s]*shot\s*learning|zero[-\s]*shot\s*learning|'
            r'online\s*learning|active\s*learning|reinforcement\s*learning|'
            r'q[-\s]*learning|deep\s*q[-\s]*network|dqn|policy\s*gradient|actor[-\s]*critic|'
            r'ppo|sac|td\s*learning|multi[-\s]*agent\s*reinforcement\s*learning|marl|'
            r'linear\s*regression|logistic\s*regression|decision\s*tree|random\s*forest|svm|'
            r'support\s*vector\s*machine|knn|k[-\s]*nearest\s*neighbors?|naive\s*bayes|'
            r'bayesian\s*network|hidden\s*markov\s*model|hmm|markov\s*decision\s*process|mdp|'
            r'gaussian\s*process|gp|k[-\s]*means|hierarchical\s*clustering|dbscan|pca|'
            r'principal\s*component\s*analysis|lda|expectation[-\s]*maximization|'
            r'em\s*algorithm|boosting|adaboost|xgboost|lightgbm|catboost|'
            r'optimization|genetic\s*algorithm|evolutionary\s*algorithm|metaheuristic|'
            r'simulated\s*annealing|particle\s*swarm\s*optimization|pso|ant\s*colony|'
            r'tabu\s*search|bayesian\s*optimization|gradient\s*descent|'
            r'stochastic\s*gradient\s*descent|sgd|adam|rmsprop|'
            r'anomaly\s*detection|outlier\s*detections?|dimensionality\s*reduction|'
            r'feature\s*engineering|feature\s*extraction|pattern\s*recognition|'
            r'simulation|simulator|model[-\s]*based|agent[-\s]*based\s*model|abm|'
            r'finite[\s\-–—]*element([\s\-–—]*method)?|fem|'
            r'forecasting|time[-\s]*series|information\s*retrieval|knowledge\s*discovery|'
            r'data\s*mining|predictive\s*model|'
            r'life[-\s]*cycle\s*assessment|lca'
            r')\b',
            re.IGNORECASE
        )


        # --- Manufacturing and production terms ---
        self.manufacturing_core = re.compile(
            r'\b(manufacturing|production\s+(system|line|planning|scheduling)|'
            r'factory|shop\s*floor|process\s+optimization|job\s*shop|flow\s*shop|'
            #r'assembly\s+line|machining|industrial\s+automation|'
            r'high[- ]mix|low[- ]volume|work\s*cell)\b',
            re.IGNORECASE
        )

        # --- Decision support focus terms ---
        self.decision_support = re.compile(
            r'\b(decision\s+support|decision\s+making|multi[- ]criteria|mcdm|'
            r'dashboard|planning\s+support|data\s+driven\s+decision|decision[- ]making|'
            r'intelligent\s+decision|decision[- ]support system|decision-support framework)\b',
            re.IGNORECASE
        )


        self.scheduling_core = re.compile(
            r'\b(job\s*shop|flow\s*shop|flexible\s*job\s*shop|fjsp|dfjsp|'
            r'dynamic\s*scheduling|dispatching\s*rules?|pdrs?|agv|dispatching|'
            r'lot[-\s]*sizing|allocation|workload\s*balancing|'
            r'production\s+plan|scheduling|sequencing|forecasting|planning|'
            r'model\s+predictive\s+control)\b',
            re.IGNORECASE
        )

        self.metaheuristics_advanced = re.compile(
            r'\b('
            r'lns|large\s*neighborhood\s*search|'
            r'vns|variable\s*neighborhood\s*search|'
            r'grasp|greedy\s*randomized\s*adaptive\s*search\s*procedure|'
            r'memetic(\s*algorithm)?|'
            r'ica|imperialist\s*competitive(\s*algorithm)?|'
            r'nsga[-\s]*ii|nsga2|non[-\s]*dominated\s*sorting\s*ga|'
            r'moea|multi[-\s]*objective\s*evolutionary\s*algorithm|'
            r'meta[-\s]*learning|hyper[-\s]*heuristic'
            r')\b',
            re.IGNORECASE
        )

        self.robotics_core = re.compile(
            r'\b(human[-\s]*robot\s*collaboration|hrc|cobot|cobots|'
            r'robotic\s*cell|robotic\s*scheduling|autonomous\s*vehicle|agv)\b',
            re.IGNORECASE
        )

        self.simulation_core = re.compile(
            r'\b(simulation[-\s]*based|simulation\s*optimization|discrete[-\s]*event\s*simulation|'
            r'des|monte\s*carlo|mcts|petri\s*net|object[-\s]*oriented\s*modeling|'
            r'simulation[-\s]*model|process\s*simulation)\b',
            re.IGNORECASE
        )

        self.monitoring_core = re.compile(
            r'\b(real[-\s]*time\s*monitoring|online\s*monitoring|'
            r'condition\s*monitoring|monitoring)\b',
            re.IGNORECASE
        )


        # --- Broader Industry 4.0 context ---
        self.industry_broad = re.compile(
            r'\b(industry\s*4\.0|industry\s*5\.0|smart\s+factory|digital\s+manufacturing|'
            r'smart\s+production|digital\s+transformation|'
            r'cloud\s*computing|big\s*data)\b',
            re.IGNORECASE
        )

        # --- Negative indicators (exclude) ---
        self.negative_indicators = re.compile(
            r'\b(business\s+model|marketing|leadership|management|finance|'
            r'human\s+resource|organization|corporate|'
            r'social\s+impact|circular\s+economy|entrepreneurship|life cycle assessment|lca|'
            r'supply\s+chain\s+finance|economic\s+policy)\b',
            re.IGNORECASE
        )

        # --- Identify manufacturing level factory ---
        self.factory_level = re.compile(
            r'\b(shop[- ]floor|flow[- ]shop|job[- ]shop)\b',
            re.IGNORECASE
        )

    # ------------------------------
    # Helper functions
    # ------------------------------
    def _count_matches(self, pattern: re.Pattern, text: str) -> int:
        if not text or pd.isna(text):
            return 0
        return len(pattern.findall(str(text))) # unique matches only or len(pattern.findall(str(text)))

    def _check_match(self, pattern: re.Pattern, text: str) -> bool:
        if not text or pd.isna(text):
            return False
        return bool(pattern.search(str(text)))

    def _get_combined_text(self, row: pd.Series) -> str:
        parts = []
        for field in ["Title", "Abstract", "Author Keywords", "Index Keywords"]:
            if field in row and pd.notna(row[field]):
                parts.append(str(row[field]))
        return " ".join(parts).lower()
    
    def clean(self, s: str) -> str:
        s = str(s)
        s = s.strip()
        s = re.sub(r"\s+", "_", s)  # replace any whitespace with _
        return s.lower()

    # ------------------------------
    # Scoring and classification logic
    # ------------------------------
    def _score_paper(self, row: pd.Series) -> Tuple[str, str, str]:
        text = self._get_combined_text(row)
        title = str(row.get("Title", "")).lower()

        # Count matches
        tech_count = self._count_matches(self.tech_core, text)
        mfg_count = self._count_matches(self.manufacturing_core, text)
        decision_count = self._count_matches(self.decision_support, text)
        industry_count = self._count_matches(self.industry_broad, text)
        robotics_count = self._count_matches(self.robotics_core, text)
        simulation_count = self._count_matches(self.simulation_core, text)
        scheduling_count = self._count_matches(self.scheduling_core, text)
        metaheuristic_count = self._count_matches(self.metaheuristics_advanced, text)
        monitoring_count = self._count_matches(self.monitoring_core, text)

        combined_count = scheduling_count + metaheuristic_count + simulation_count + robotics_count + monitoring_count

        has_negative = self._check_match(self.negative_indicators, text)

        # Logic tree (same structure as original)
        #if has_negative:
        #    return "-", "Managerial/business/policy focus"

        if mfg_count == 0:
            return "-", "No manufacturing/production focus", "Managerial/business/policy focus" if has_negative else "None"

        # ++ Most relevant old
        #if tech_count >= 2 and decision_count >= 1 and mfg_count >= 1:
        #    return "++", f"Core tech ({tech_count}) + decision support ({decision_count}) + manufacturing ({mfg_count})"

        # ++ Most relevant new
        #if tech_count >= 2 and mfg_count >= 1 and (decision_count >= 2 or combined_count >= 3):
        if tech_count >= 2 and mfg_count >= 1 and (decision_count >= 1 or combined_count >= 2): #tech_count >= 2 and mfg_count >= 1 and (decision_count >= 1 or combined_count >= 2):
            return "++", f"Core tech ({tech_count}) + manufacturing ({mfg_count}) + strong decision or scheduling/simulation/robotics signal", "Managerial/business/policy focus" if has_negative else "None"

        # + Relevant
        if tech_count >= 2 and combined_count >= 1 and mfg_count >= 1: # tech_count >= 2 and combined_count >= 2 and mfg_count >= 1:
            return "+", f"Tech ({tech_count}) + scheduling/metaheuristics/simulation/robotics ({combined_count}) in manufacturing", "Managerial/business/policy focus" if has_negative else "None"
        
        if industry_count >= 2 and tech_count >= 1 and mfg_count >= 1: # industry_count >= 2 and tech_count >= 1 and mfg_count >= 1
            return "+", f"Industry 4.0 digitalization ({industry_count}) + tech ({tech_count})", "Managerial/business/policy focus" if has_negative else "None"

        # * Maybe
        if tech_count >= 1 and mfg_count >= 1:
            return "*", f"Manufacturing ({mfg_count}) + some technology relevance ({tech_count})", "Managerial/business/policy focus" if has_negative else "None"

        # - Exclude
        return "-", "Not relevant to decision support or Industry 4.0 manufacturing", "Managerial/business/policy focus" if has_negative else "None"

    # ------------------------------
    # Main processing loop
    # ------------------------------
    def process_papers(self, input_csv: str, output_excel: str, sheet_name: str = "Automated Selection") -> None:
        print(f"Loading papers from {input_csv}...")
        df = pd.read_csv(input_csv, sep=",", encoding="utf-8")
        print(f"Loaded {len(df)} papers")

        results = []
        for i, (_, row) in enumerate(df.iterrows()):
            keep_decision, reasoning, negative = self._score_paper(row)
            methods_found = [m.group(0) for m in self.methods_core.finditer(self._get_combined_text(row))]
            tech_found = [m.group(0) for m in self.tech_core.finditer(self._get_combined_text(row))]
            if keep_decision == "*" or keep_decision == "-":
                continue  # To get only high and relevant papers
            results.append({
                "key_id": "_".join([self.clean(row.get("Year","")), self.clean(row.get("Title",""))]), #row.get("key_id", ""),
                "Title": row.get("Title", ""),
                "Relevance Score": keep_decision,
                "Reasoning": reasoning,
                "Methods": ", ".join(sorted(set(methods_found))),
                "Technologies": ", ".join(sorted(set(tech_found))),
                "Automated Manufacturing Level": "Factory" if self._count_matches(self.factory_level, self._get_combined_text(row)) else "Unclassified",
                "Negatives Found": negative,
                "Include/Exclude": "",
                "Your Notes": "ex. reason for inclusion/exclusion",
                "EID": row.get("EID", ""), # For citation fetching later
                "Year": row.get("Year", ""), # For citation fetching later
            })
            if (i + 1) % 100 == 0:
                print(f"Processed {i + 1}/{len(df)} papers...")

        output_df = pd.DataFrame(results)

        # If csv target, write to CSV
        if output_excel.lower().endswith(".csv"):
            output_df.to_csv(output_excel, index=False)
            print(f"\n✅ Results written to '{output_excel}'.")
        
        else:
            # Write to Excel (preserve other sheets)
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.worksheet.table import Table, TableStyleInfo

            try:
                wb = load_workbook(output_excel)
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                ws = wb.create_sheet(sheet_name, 10)
                for r in dataframe_to_rows(output_df, index=False, header=True):
                    ws.append(r)
                last_col, last_row = ws.max_column, ws.max_row
                ref = f"A1:{ws.cell(row=last_row, column=last_col).coordinate}"
                tab = Table(displayName="PaperSelection", ref=ref)
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                tab.tableStyleInfo = style
                ws.add_table(tab)
                wb.save(output_excel)
                print(f"\n✅ Results written to '{output_excel}' sheet '{sheet_name}'.")
            except FileNotFoundError:
                with pd.ExcelWriter(output_excel, engine="openpyxl", mode="w") as writer:
                    output_df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"\n✅ Created new '{output_excel}' with sheet '{sheet_name}'.")

        # Summary
        summary = output_df["Relevance Score"].value_counts().sort_index()
        print("\nSummary:")
        print(f"  Total papers: {len(output_df)}")
        for decision in ["++", "+", "*", "-"]: #, "~",
            count = summary.get(decision, 0)
            pct = (count / len(output_df) * 100) if len(output_df) > 0 else 0
            print(f"  {decision}: {count} ({pct:.1f}%)")

        # Print sample of most relevant papers
        top = output_df[output_df["Relevance Score"] == "++"].head(5)
        if not top.empty:
            print("\nSample of most relevant papers:")
            for _, r in top.iterrows():
                print(f"  - {r['key_id'][:80]}...")
                print(f"    Reason: {r['Reasoning']}")


def main():
    """Main entry point."""
    config, repo_root = load_paths_config()
    defaults = config["defaults"]["relevance_score"]

    selector = PaperSelector()
    input_file = resolve_from_root(repo_root, defaults["input_file"])
    output_excel = resolve_from_root(repo_root, defaults["output_file"])
    sheet_name = defaults["sheet_name"]
    selector.process_papers(input_file, output_excel, sheet_name)
    print("\nAutomated HMLV paper selection completed!")
    print(f"Review the results in '{output_excel}' sheet '{sheet_name}'.")


if __name__ == "__main__":
    main()
