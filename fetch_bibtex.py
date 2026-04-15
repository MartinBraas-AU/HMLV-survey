"""
Fetch BibTeX entries for missing references in the Cleaned Master Sheet.
Uses CrossRef content negotiation via DOI — no API key needed.

Cite keys are formatted as AuthorYearFirstword (e.g. Salatiello2024Longsighted).

Usage:
    pip install openpyxl pandas requests
    python fetch_bibtex.py Master_sheet.xlsx          # fetch missing + update sheet + .bib
    python fetch_bibtex.py --rekey existing.bib        # rekey an existing .bib file in-place

Outputs:
    - Updated Master_sheet.xlsx with BibTeX keys in 'Bibtex ref' column
    - missing_refs.bib with full BibTeX entries
"""

import sys
import re
import time
import pandas as pd
from openpyxl import load_workbook
import requests

SHEET_NAME = "Cleaned Master sheet"
BIBTEX_COL = "Bibtex ref"
DOI_COL = "DOI"


def fetch_bibtex(doi: str) -> str | None:
    """Fetch BibTeX string from doi.org via content negotiation."""
    url = f"https://doi.org/{doi}"
    headers = {"Accept": "application/x-bibtex"}
    try:
        r = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        r.raise_for_status()
        return r.text.strip()
    except Exception as e:
        print(f"  FAIL {doi}: {e}")
        return None


def extract_cite_key(bibtex: str) -> str | None:
    """Extract the citation key from a BibTeX entry, e.g. @article{Smith2024, ..."""
    m = re.match(r"@\w+\{(.+?),", bibtex)
    return m.group(1) if m else None


def extract_field(bibtex: str, field: str) -> str | None:
    """Extract a field value from a BibTeX entry."""
    pattern = rf"{field}\s*=\s*\{{(.+?)\}}"
    m = re.search(pattern, bibtex, re.IGNORECASE | re.DOTALL)
    return m.group(1).strip() if m else None


def make_cite_key(bibtex: str) -> str | None:
    """Generate AuthorYearFirstword cite key, e.g. Salatiello2024Longsighted."""
    # Extract author — take first author's surname
    author = extract_field(bibtex, "author")
    if not author:
        return None
    # First author is before the first " and " or ","
    first_author = re.split(r"\s+and\s+|,", author)[0].strip()
    # Surname is typically the last word (or first if "Surname, Firstname" format)
    # If comma-separated (Surname, First), take before comma
    if "," in first_author.split(" and ")[0]:
        surname = first_author.split(",")[0].strip()
    else:
        # "First Last" format — take last word
        surname = first_author.split()[-1].strip()
    # Clean LaTeX accents and non-alpha chars
    surname = re.sub(r"[{}\\\"]", "", surname)
    surname = surname.capitalize()

    # Extract year
    year = extract_field(bibtex, "year")
    if not year:
        # Try month/date patterns or doi for year
        year = "XXXX"
    year = re.sub(r"[^0-9]", "", year)[:4]

    # Extract title — take first meaningful word (skip articles)
    title = extract_field(bibtex, "title")
    if not title:
        return f"{surname}{year}"
    # Remove LaTeX braces/commands
    title_clean = re.sub(r"[{}\\]", "", title)
    words = title_clean.split()
    skip = {"a", "an", "the", "on", "of", "for", "in", "to", "and", "with", "from", "by", "towards"}
    first_word = "Unknown"
    for w in words:
        w_alpha = re.sub(r"[^a-zA-Z]", "", w)
        if w_alpha.lower() not in skip and len(w_alpha) > 1:
            first_word = w_alpha.capitalize()
            break

    return f"{surname}{year}{first_word}"


def rewrite_cite_key(bibtex: str, new_key: str) -> str:
    """Replace the cite key in a BibTeX entry."""
    return re.sub(r"(@\w+\{).+?,", rf"\g<1>{new_key},", bibtex, count=1)


def deduplicate_key(key: str, seen: set) -> str:
    """Append a/b/c suffix if key already used."""
    if key not in seen:
        seen.add(key)
        return key
    for suffix in "abcdefghijklmnopqrstuvwxyz":
        candidate = f"{key}{suffix}"
        if candidate not in seen:
            seen.add(candidate)
            return candidate
    return key  # shouldn't happen with 26 suffixes


def rekey_bibfile(bib_path: str):
    """Rekey all entries in an existing .bib file to AuthorYearFirstword format."""
    with open(bib_path, "r", encoding="utf-8") as f:
        content = f.read()

    entries = re.split(r"\n(?=@)", content)
    seen_keys: set[str] = set()
    new_entries = []
    for entry in entries:
        entry = entry.strip()
        if not entry:
            continue
        new_key = make_cite_key(entry)
        if new_key:
            new_key = deduplicate_key(new_key, seen_keys)
            entry = rewrite_cite_key(entry, new_key)
            print(f"  {new_key}")
        new_entries.append(entry)

    with open(bib_path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(new_entries) + "\n")
    print(f"\nRekeyed {len(new_entries)} entries in {bib_path}")


def main(xlsx_path: str):
    # --- Read with pandas to find missing rows ---
    df = pd.read_excel(xlsx_path, sheet_name=SHEET_NAME)
    missing_mask = df[BIBTEX_COL].isna() | (df[BIBTEX_COL].astype(str).str.strip() == "")
    has_doi_mask = df[DOI_COL].notna() & (df[DOI_COL].astype(str).str.strip() != "") & (df[DOI_COL].astype(str).str.strip() != "-")
    targets = df[missing_mask & has_doi_mask]

    print(f"Found {len(targets)} entries missing BibTeX with a DOI.\n")

    # --- Fetch BibTeX for each DOI ---
    bib_entries = []  # (dataframe_index, cite_key, full_bibtex)
    seen_keys: set[str] = set()
    for idx, row in targets.iterrows():
        doi_raw = str(row[DOI_COL]).strip()
        doi = doi_raw.split("doi.org/")[-1] if "doi.org/" in doi_raw else doi_raw
        title_short = str(row["Title"])[:70]

        print(f"[{len(bib_entries)+1}/{len(targets)}] {title_short}...")
        bibtex = fetch_bibtex(doi)
        if bibtex:
            new_key = make_cite_key(bibtex)
            if new_key:
                new_key = deduplicate_key(new_key, seen_keys)
                bibtex = rewrite_cite_key(bibtex, new_key)
            else:
                new_key = extract_cite_key(bibtex)  # fallback to original
            bib_entries.append((idx, new_key, bibtex))
            print(f"  OK -> {new_key}")
        time.sleep(0.5)

    # --- Write .bib file ---
    bib_path = "missing_refs.bib"
    with open(bib_path, "w", encoding="utf-8") as f:
        for _, _, bibtex in bib_entries:
            f.write(bibtex + "\n\n")
    print(f"\nWrote {len(bib_entries)} entries to {bib_path}")

    # --- Update spreadsheet ---
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]

    # Find column index for 'Bibtex ref' (1-indexed)
    header_row = [cell.value for cell in ws[1]]
    bib_col_idx = header_row.index(BIBTEX_COL) + 1  # openpyxl is 1-indexed

    updated = 0
    for df_idx, cite_key, _ in bib_entries:
        if cite_key:
            excel_row = df_idx + 2  # +1 for header, +1 for 1-indexing
            ws.cell(row=excel_row, column=bib_col_idx, value=cite_key)
            updated += 1

    wb.save(xlsx_path)
    print(f"Updated {updated} cells in '{SHEET_NAME}' -> '{BIBTEX_COL}' column.")
    print("Done!")


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--rekey":
        rekey_bibfile(sys.argv[2])
    else:
        path = sys.argv[1] if len(sys.argv) > 1 else "data/Master sheet.xlsx"
        main(path)